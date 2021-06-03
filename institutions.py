from numpy import printoptions
from openpyxl import load_workbook
import pandas as pd

datasheet = "synthesio.xlsx"
synthesio = pd.read_excel(datasheet, sheet_name="export")

comments = synthesio["Mention Content"]
topics = synthesio["Topics"]
gscore = synthesio["Google Sentiment Score"]
gmagnitude = synthesio['Google Sentiment Magnitude']
iname = ["Institution Name"]
ipositive = ["Número de comentarios positivos"]
inegative = ["Número de comentarios negativos"]
ineutral = ["Número de comentarios neutrales"]
igrade = ["Sumatoria Goggle Sentiment Score"]
iaverage = ["Promedio"]
isentiment = ["Sentimiento final"]
institutions = []


class RdInstitution:
    def __init__(self, name, positive, negative, neutral, grade, average, sentiment, com):
        self.name = name
        self.positive = positive
        self.negative = negative
        self.neutral = neutral
        self.grade = grade
        self.average = average
        self.sentiment = sentiment
        self.com = com


for line in topics:
    text = line.split(",")
    for x in text:
        count = 0
        for a in institutions:
            if x.strip() == a.name:
                count = count + 1
        if count == 0:
            institutions.append(RdInstitution(x.strip(), 0, 0, 0, 0, 0, 0, []))

currentLine = 0
for line in topics:
    text = line.split(",")
    for x in text:
        for inst in institutions:
            if inst.name == x.strip():
                inst.com.append(comments[currentLine])
                if gscore[currentLine] > 0:
                    inst.positive = inst.positive + 1
                elif gscore[currentLine] < 0:
                    inst.negative = inst.negative + 1
                else:
                    inst.neutral = inst.neutral + 1
                inst.grade = inst.grade + gscore[currentLine]
                inst.average = inst.grade / \
                    (inst.positive + inst.neutral + inst.negative)
                if inst.average < 0:
                    inst.sentiment = "Negativo"
                elif inst.average > 0:
                    inst.sentiment = "Positivo"
                else:
                    inst.sentiment = "Neutral"
    currentLine = currentLine + 1


def printer(dataList):
    dataFrame = pd.DataFrame({'Data': dataList})
    for index, row in dataFrame.iterrows():
        if dataList[0] == "Institution Name":
            cell = 'BC%d' % (index + 1)
            ws[cell] = row[0]
        elif dataList[0] == "Número de comentarios positivos":
            cell = 'BD%d' % (index + 1)
            ws[cell] = row[0]
        elif dataList[0] == "Número de comentarios negativos":
            cell = 'BE%d' % (index + 1)
            ws[cell] = row[0]
        elif dataList[0] == "Número de comentarios neutrales":
            cell = 'BF%d' % (index + 1)
            ws[cell] = row[0]
        elif dataList[0] == "Sumatoria Goggle Sentiment Score":
            cell = 'BG%d' % (index + 1)
            ws[cell] = row[0]
        elif dataList[0] == "Promedio":
            cell = 'BH%d' % (index + 1)
            ws[cell] = row[0]
        else:
            cell = 'BI%d' % (index + 1)
            ws[cell] = row[0]


def framer(insititutionsObj):
    for inst in insititutionsObj:
        iname.append(inst.name)
        ipositive.append(inst.positive)
        inegative.append(inst.negative)
        ineutral.append(inst.neutral)
        igrade.append(inst.grade)
        iaverage.append(inst.average)
        isentiment.append(inst.sentiment)
    printer(iname)
    printer(ipositive)
    printer(inegative)
    printer(ineutral)
    printer(igrade)
    printer(iaverage)
    printer(isentiment)


wb = load_workbook(datasheet)
ws = wb['export']
framer(institutions)
wb.save(datasheet)