# Imports the Google Cloud client library

import ctypes  # An included library with Python install.
from openpyxl import load_workbook
import pandas as pd
from google.cloud import language_v1
from google.oauth2 import service_account
from pyasn1.type.constraint import ConstraintsUnion

# ctypes.windll.user32.MessageBoxW(
    # 0, "Algoritmo en ejecución.", "Por favor espera un momento", 1)

credentials = service_account.Credentials.from_service_account_file(
    "key.json")

# Instantiates a client
client = language_v1.LanguageServiceClient(credentials=credentials)

# The text to analyze

datasheet = "synthesio.xlsx"

synthesio = pd.read_excel(datasheet, sheet_name="export")

comments = synthesio["Mention Content"]
gscore = ['Google Sentiment Score']
gmagnitude = ['Google Sentiment Magnitude']
gsentiment = ['Google Sentiment']

for line in comments:
    text = line
    document = language_v1.Document(
        content=text, type_=language_v1.Document.Type.PLAIN_TEXT, language='es')
    sentiment = client.analyze_sentiment(
        request={'document': document}).document_sentiment
    if sentiment.score > 0:
        sentimentValue = "Positivo"
    elif sentiment.score < 0:
        sentimentValue = "Negativo"
    else:
        sentimentValue = "Neutral"
    gscore.append(sentiment.score)
    gmagnitude.append(sentiment.magnitude)
    gsentiment.append(sentimentValue)

gScoreResults = pd.DataFrame({'Data': gscore})
gMagnitudeResults = pd.DataFrame({'Data': gmagnitude})
gSentimentResults = pd.DataFrame({'Data': gsentiment})

wb = load_workbook(datasheet)

ws = wb['export']

for index, row in gScoreResults.iterrows():
    cell = 'AZ%d' % (index + 1)
    ws[cell] = row[0]

for index, row in gMagnitudeResults.iterrows():
    cell = 'BA%d' % (index + 1)
    ws[cell] = row[0]

for index, row in gSentimentResults.iterrows():
    cell = 'BB%d' % (index + 1)
    ws[cell] = row[0]

wb.save(datasheet)

import institutions

# ctypes.windll.user32.MessageBoxW(
    # 0, "El algoritmo se ejecutó exitosamente. Revisa el archivo de Excel synthesio.xlsx", "Resultado Exitoso", 1)